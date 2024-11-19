from io import BytesIO
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import PatternFill, Font
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from rest_framework.response import Response
from Utils.ResposeData import ResponseData
from rest_framework import viewsets, status, serializers, response
from rest_framework.decorators import action
from Apps.Catalogos.empleados.models import Empleado
from Apps.Catalogos.empleados.API.serializers import EmpleadoSerializer

class EmpleadoViewSet(viewsets.ModelViewSet):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer
    # Sobreescribir el metodo create
    def create(self, request, *args, **kwargs):
        numero_cedula = request.data.get('numero_cedula')

        # Verificar si ya existe ese numero de cédula
        if Empleado.objects.filter(numero_cedula=numero_cedula, is_active=True).exists():
            response_data = ResponseData(
                Success=False,
                Status=status.HTTP_400_BAD_REQUEST,
                Message='Ya existe un empleado registrado con este número de cédula.',
                Record=None
            )
            return Response(response_data.toResponse(), status=status.HTTP_400_BAD_REQUEST)

        #Serializar datos
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        response_data = ResponseData(
            Success=True,
            Status=status.HTTP_201_CREATED,
            Message='Empleado creado con éxito.',
            Record=serializer.data
        )
        return Response(response_data.toResponse(), status=status.HTTP_201_CREATED)

    # Sobreescribir list
    def list(self, request, *args, **kwargs):
        numero_cedula = request.data.get('numero_cedula') or request.GET.get('numero_cedula')

        if numero_cedula:
            queryset = Empleado.objects.filter(numero_cedula=numero_cedula, is_active= True)
            if not queryset.exists():
                response_data = ResponseData(
                    Success=False,
                    Status=status.HTTP_404_NOT_FOUND,
                    Message=f"No hay un empleado asociado a esta cédula de identidad: {numero_cedula}",
                    Record=None
                )
                return Response(response_data.toResponse(), status=status.HTTP_404_NOT_FOUND)
            else:
                message = f"Empleado encontrado con: {numero_cedula}"
        else:
            queryset = Empleado.objects.filter(is_active= True)
            message = "Lista de empleados activos"

        serializer = self.get_serializer(queryset, many=True)
        data = ResponseData(
            Success=True,
            Status=status.HTTP_200_OK,
            Message=message,
            Record=serializer.data
        )
        return Response(data.toResponse(), status=status.HTTP_200_OK)

    # Eliminar por número de cédula
    @action(detail=False, methods=['delete'], url_path='eliminar-cedula')
    def eliminar_cedula(self, request):
        numero_cedula = request.data.get('numero_cedula')
        if not numero_cedula:
            data = ResponseData(
                Success=False,
                Status=status.HTTP_400_BAD_REQUEST,
                Message="Digite el número de cédula.",
                Record=None
            )
            return Response(data.toResponse(), status=status.HTTP_400_BAD_REQUEST)

        empleado = Empleado.objects.filter(numero_cedula=numero_cedula, is_active=True).first()

        if not empleado:
            response_data = ResponseData(
                Success=False,
                Status=status.HTTP_400_BAD_REQUEST,
                Message=f'No hay un empleado asociado a este número de cédula: {numero_cedula}',
                Record=None
            )
            return Response(response_data.toResponse(), status=status.HTTP_400_BAD_REQUEST)

        empleado.is_active = False
        empleado.save()
        serializer = self.get_serializer(empleado)

        response_data = ResponseData(
            Success=True,
            Status=status.HTTP_200_OK,
            Message='Empleado eliminado con éxito.',
            Record=serializer.data
        )
        return Response(response_data.toResponse(), status=status.HTTP_200_OK)


    # Actualizar por número de cédula
    @action(detail=False, methods=['put'], url_path='actualizar-empleado-cedula')
    def actualizar_empleado_cedula(self, request):
        numero_cedula = request.data.get('numero_cedula')

        if not numero_cedula:
            data = ResponseData(
                Success=False,
                Status=status.HTTP_400_BAD_REQUEST,
                Message="Digite el número de cédula.",
                Record=None
            )
            return Response(data.toResponse(), status=status.HTTP_400_BAD_REQUEST)

        empleado = Empleado.objects.filter(numero_cedula=numero_cedula, is_active=True).first()
        if not empleado:
            response_data = ResponseData(
                Success=False,
                Status=status.HTTP_404_NOT_FOUND,
                Message='Empleado no encontrado',
                Record=None
            )
            return Response(response_data.toResponse(), status=status.HTTP_404_NOT_FOUND)

        serializer = EmpleadoSerializer(empleado, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            response_data = ResponseData(
                Success=True,
                Status=status.HTTP_200_OK,
                Message='Empleado actualizado con éxito',
                Record=serializer.data
            )
            return Response(response_data.toResponse(), status=status.HTTP_200_OK)

        response_data = ResponseData(
            Success=False,
            Status=status.HTTP_400_BAD_REQUEST,
            Message='Error en la validación',
            Record=serializer.errors
        )
        return Response(response_data.toResponse(), status=status.HTTP_400_BAD_REQUEST)

    # CAMBIAR NUMERO DE CEDULA
    @action(detail=False, methods=['patch'], url_path='actualizar-cedula')
    def actualizar_cedula(self, request):
        cedula_actual = request.data.get('cedula_actual')
        nueva_cedula = request.data.get('nueva_cedula')

        if not cedula_actual or not nueva_cedula:
            data = ResponseData(
                Success=False,
                Status=status.HTTP_400_BAD_REQUEST,
                Message="Falta un valor, cédula registrada, o la nueva versión de la cédula",
                Record=None
            )
            return Response(data.toResponse(), status=status.HTTP_400_BAD_REQUEST)

        if Empleado.objects.filter(numero_cedula=nueva_cedula, is_active=True).exclude(
                numero_cedula=cedula_actual).exists():
            data = ResponseData(
                Success=False,
                Status=status.HTTP_400_BAD_REQUEST,
                Message=f"El número de cédula: {nueva_cedula} esta asociada a otro empleado, verifique nuevamente",
                Record=None
            )
            return Response(data.toResponse(), status=status.HTTP_400_BAD_REQUEST)

        empleado = Empleado.objects.filter(numero_cedula=cedula_actual, is_active=True).first()
        if not empleado:
            response_data = ResponseData(
                Success=False,
                Status=status.HTTP_404_NOT_FOUND,
                Message=f"No hay un empleado con este número de cédula: {cedula_actual}",
                Record=None
            )
            return Response(response_data.toResponse(), status=status.HTTP_404_NOT_FOUND)

        empleado.numero_cedula = nueva_cedula
        empleado.save()
        serializer = self.get_serializer(empleado)
        data = ResponseData(
            Success=True,
            Status=status.HTTP_200_OK,
            Message="Número de cédula actualizado con éxito",
            Record=serializer.data
        )
        return Response(data.toResponse(), status=status.HTTP_200_OK)


    #TEST OF GENERATION OF REPORT IN PDF AND EXCEL
    #EXCEL
    @action(detail=False, methods=['post'], url_path='reporte')
    def generar_reporte(self, request):
        sexo = request.data.get("sexo") or request.GET.get("sexo")
        if sexo:
            empleados = Empleado.objects.filter(sexo=sexo, is_active=True)
            if not empleados:
                response_data = ResponseData(
                    Success=False,
                    Status=status.HTTP_404_NOT_FOUND,
                    Message="No hay empleados con ese género",
                    Record=None
                )
                return Response(response_data.toResponse(), status=status.HTTP_404_NOT_FOUND)
        else:
            empleados = Empleado.objects.filter(is_active=True)

        #creacion del libro
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = "Registros de empleados por género"

        encabezados = ["ID", "Cédula", "Inss", "Primer nombre", "Segundo nombre", "Primer apellido", "Segundo apellido",
                       "Dirección", "Sexo", "Télefono", "Email", "Estado"]

        # personalizacion
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        font = Font(bold=True)

        for col_num, encabezado in enumerate(encabezados, start=1):
            cell = hoja.cell(row=1, column=col_num, value=encabezado)
            cell.fill = fill
            cell.font = font

        for empleado in empleados:
            hoja.append([empleado.id_empleados, empleado.numero_cedula, empleado.numero_inss, empleado.primer_nombre,
                         empleado.segundo_nombre,
                         empleado.primer_apellido, empleado.segundo_apellido, empleado.direccion, empleado.sexo,
                         empleado.telefono,
                         empleado.email, empleado.estado])

        for col in hoja.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            hoja.column_dimensions[column].width = adjusted_width

       #RESPUESTA WITH THE EXCEL
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Empleados_Lista.xlsx'

        workbook.save(response)
        return response

    # intento mil de un pdf:(
    @action(detail=False, methods=['post'], url_path='reporte-pdf')
    def generar_reporte_pdf(self, request):
        sexo = request.data.get("sexo") or request.GET.get("sexo")

        if sexo:
            empleados = Empleado.objects.filter(sexo=sexo, is_active=True)
            if not empleados:
                response_data = ResponseData(
                    Success=False,
                    Status=status.HTTP_404_NOT_FOUND,
                    Message="No hay empleados de ese género",
                    Record=None
                )
                return Response(response_data.toResponse(), status=status.HTTP_404_NOT_FOUND)
        else:
            empleados = Empleado.objects.filter(is_active=True)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)

        encabezados = ["ID", "Cédula", "Inss", "Nombre1", "Nombre2", "Apellido1", "Apellido2",
                       "Dirección", "Sexo", "Télefono", "Email", "Estado"]

        data = [encabezados]
        for empleado in empleados:
            data.append([
                str(empleado.id_empleados),
                str(empleado.numero_cedula),
                str(empleado.numero_inss),
                str(empleado.primer_nombre),
                str(empleado.segundo_nombre),
                str(empleado.primer_apellido),
                str(empleado.segundo_apellido),
                str(empleado.direccion),
                str(empleado.sexo),
                str(empleado.telefono),
                str(empleado.email),
                str(empleado.estado)
            ])
        table = Table(data)

        # personalizacion de la tabla
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), (0.5, 0.5, 0.5)),
            ('TEXTCOLOR', (0, 0), (-1, 0), (1, 1, 1)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), (1, 1, 1)),
            ('GRID', (0, 0), (-1, -1), 0.5, (0, 0, 0)),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))

        elements = [table]

        def agregar_encabezado(canvas, doc):
            canvas.setFont("Helvetica-Bold", 16)
            canvas.drawString(200, 750, "Lista de empleados")

            canvas.setFont("Helvetica", 12)
            canvas.drawString(200, 730,
                              "Reporte generado para el género: {}".format(
                                  sexo if sexo else None))

            canvas.drawString(200, 710, " ")

        doc.build(elements, onFirstPage=agregar_encabezado)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=Lista_de_empleados.pdf'

        return response
